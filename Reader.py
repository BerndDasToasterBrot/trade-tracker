import os
import re
import pandas as pd
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from datetime import datetime

# Define paths
PDF_FOLDER = os.path.join(os.path.dirname(__file__), 'pdfs')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'Trading.xlsx')

def parse_number(text):
    """Wandelt Strings wie '1.10', '0,06', '72.00' in float um."""
    if not text:
        return 0.0
    text = text.replace('EUR', '').replace('€', '')
    text = text.replace('pc.', '').replace('pc', '')
    text = text.strip()
    # Deutsch: 1.234,56  -> 1234.56
    if ',' in text and '.' in text and text.find('.') < text.find(','):
        text = text.replace('.', '').replace(',', '.')
    else:
        text = text.replace(',', '.')
    return float(text)


def normalize_asset_name(name):
    """Macht aus dem Titel eine Token-Menge, um fuzzy zu matchen."""
    if not isinstance(name, str):
        return set()
    name = name.lower()
    # offensichtlichen Kram raus
    for ch in ',.$€':
        name = name.replace(ch, ' ')
    cleaned = ''.join(
        ch if (ch.isalnum() or ch.isspace()) else ' '
        for ch in name
    )
    tokens = [t for t in cleaned.split() if t]
    return set(tokens)


def find_best_matching_row(sheet, asset_name, require_empty_sell_date=True, min_score=0.5):
    """
    Sucht die Zeile mit dem besten Token-Overlap.
    require_empty_sell_date=True: nur Zeilen ohne Sell-Datum (für Cost-Info-Sells)
    require_empty_sell_date=False: auch Closed Trades (für Contract-Notes mit Steuern).
    """
    target_tokens = normalize_asset_name(asset_name)
    best_row = None
    best_score = 0.0

    for row in range(2, sheet.max_row + 1):
        if require_empty_sell_date and sheet[f'F{row}'].value:
            # schon verkauft -> für Cost-Info-Sell ignorieren
            continue

        row_name = sheet[f'A{row}'].value
        row_tokens = normalize_asset_name(row_name)
        if not row_tokens:
            continue

        inter = len(target_tokens & row_tokens)
        union = len(target_tokens | row_tokens)
        if union == 0:
            continue

        score = inter / union
        if score > best_score:
            best_score = score
            best_row = row

    if best_score < min_score:
        return None
    return best_row

def extract_cost_info_trade(pdf_text):
    trade_data = {}

    # Quelle markieren
    trade_data['Source'] = 'cost_info'

    # Determine if it's a Buy or Sell
    if re.search(r'Buy', pdf_text, re.IGNORECASE):
        trade_data['Trade Type'] = 'Buy'
    elif re.search(r'Sell', pdf_text, re.IGNORECASE):
        trade_data['Trade Type'] = 'Sell'
    else:
        print("Trade type not found.")
        return None

    # Asset Name aus Ex-Ante cost information
    asset_match = re.search(r'Ex-Ante cost information\s*\n\s*(.+)', pdf_text)
    if asset_match:
        trade_data['Asset Name'] = asset_match.group(1).strip()
    else:
        print("Asset Name/ISIN not found.")
        return None

    # Date
    date_match = re.search(r'Date\s*(?:\n.*?){3}\n\s*(\d{2}\.\d{2}\.\d{4})', pdf_text)
    if date_match:
        trade_data['Date'] = date_match.group(1).strip()
    else:
        print("Date not found.")
        return None

    # Quantity
    quantity_match = re.search(r'([\d,\.]+)\s*Shr\.', pdf_text)
    if quantity_match:
        trade_data['Quantity'] = float(quantity_match.group(1).replace(',', '.').strip())
    else:
        print("Quantity not found.")
        return None

    # Price per Unit (aus Gesamtbetrag / Stückzahl)
    price_match = re.search(r'Shr\.\s*\n\s*([\d,\.]+)\s*€', pdf_text)
    if price_match:
        total_order_amount = float(price_match.group(1).replace(',', '.').strip())
        trade_data['Price per Unit'] = total_order_amount / trade_data['Quantity']
    else:
        print("Price per unit not found.")
        return None

    trade_data['Fees'] = 0.99
    print(trade_data)

    return trade_data

def extract_contract_note_trade(pdf_text):
    trade_data = {}
    trade_data['Source'] = 'contract_note'

    # --- Scalable Contract note ---
    # z.B.:
    # "Contract note 
    #  Sell 72.00 pc. Microsoft Call 478,00 BNP"
    if 'Contract note' in pdf_text:
        header_match = re.search(r'(Buy|Sell)\s+([\d\.,]+)\s*pc\.\s+(.+)', pdf_text)
        if not header_match:
            print("Could not find Buy/Sell header line in Scalable contract note.")
            return None

        trade_word, qty_str, asset_name = header_match.groups()
        trade_data['Trade Type'] = 'Buy' if trade_word.lower().startswith('buy') else 'Sell'
        trade_data['Asset Name'] = asset_name.strip()
        trade_data['Quantity'] = parse_number(qty_str)

        # Datum: erste dd.mm.yyyy im Dokument
        date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', pdf_text)
        if date_match:
            trade_data['Date'] = date_match.group(1)

        # Preis pro Stück aus "72.00 pc. 1.74 EUR"
        qp_match = re.search(r'([\d\.,]+)\s*pc\.\s*([\d\.,]+)\s*EUR', pdf_text)
        if qp_match:
            qty = parse_number(qp_match.group(1))
            price = parse_number(qp_match.group(2))
            trade_data['Quantity'] = qty
            trade_data['Price per Unit'] = price

        # Steuern: letzte 4 EUR-Beträge sind (Cap, Soli, Church, Total)
        eur_values = re.findall(r'([\d\.,]+)\s*EUR', pdf_text)
        if len(eur_values) >= 4:
            tax_vals = eur_values[-4:-1]  # 1.10, 0.06, 0.10 in deinem Beispiel
            cg = parse_number(tax_vals[0])
            soli = parse_number(tax_vals[1])
            church = parse_number(tax_vals[2])
            trade_data['Capital Gains Tax'] = cg
            trade_data['Solidarity Surcharge'] = soli
            trade_data['Church Tax'] = church
            trade_data['Taxes Sum'] = cg + soli + church

        return trade_data

    # --- Baader / WPABRECHNUNG-060.114 ---
    if 'Transaction Statement: Sale' in pdf_text or 'Transaction Statement: Purchase' in pdf_text:
        if 'Transaction Statement: Sale' in pdf_text:
            trade_data['Trade Type'] = 'Sell'
        else:
            trade_data['Trade Type'] = 'Buy'

        # Asset-Name: Block zwischen WKN: ... und "Order placed by:"
        asset_block_match = re.search(
            r'WKN:[^\n]*\n(.*?)\nOrder placed by:',
            pdf_text,
            re.S
        )
        if asset_block_match:
            block = asset_block_match.group(1)
            lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
            if not lines:
                print("Baader asset block empty.")
                return None

            # offensichtlichen Müll rausfiltern: Price/EUR/reine Zahlen
            cleaned = []
            for ln in lines:
                lower = ln.lower()
                if lower in ("price", "eur"):
                    continue
                if re.fullmatch(r'[\d\.,]+', ln):
                    continue
                cleaned.append(ln)
            if cleaned:
                lines = cleaned

            # 1) Derivate: Zeile mit "Put" oder "Call" bevorzugen
            deriv_lines = [ln for ln in lines if re.search(r'\b(put|call)\b', ln, re.I)]
            if deriv_lines:
                asset_name = deriv_lines[-1]
            else:
                # 2) Aktien / sonstiges: meist ist die erste Zeile der eigentliche Name
                asset_name = lines[0]

            trade_data['Asset Name'] = asset_name
            print(f"[DEBUG] Baader Asset Name: {asset_name}")
        else:
            print("Asset name not found in Baader contract note (block regex).")
            return None

        # (optional) Menge aus "Quantity\n\nUnits   22"
        qty_match = re.search(r'Quantity\s*\n\s*Units\s+([\d\.,]+)', pdf_text)
        if qty_match:
            trade_data['Quantity'] = parse_number(qty_match.group(1))

        # Preis pro Stück aus dem "Price / EUR / 4.672" Block
        price_match = re.search(r'Price\s*EUR\s*([\d\.,]+)', pdf_text)
        if price_match:
            trade_data['Price per Unit'] = parse_number(price_match.group(1))

        # Datum: erstes ISO-Datum direkt nach "Transaction Statement: ..."
        date_match = re.search(
            r'Transaction Statement: (?:Sale|Purchase)\s*\n\s*(\d{4}-\d{2}-\d{2})',
            pdf_text
        )
        if date_match:
            iso_date = date_match.group(1)
            dt = datetime.strptime(iso_date, "%Y-%m-%d")
            trade_data['Date'] = dt.strftime("%d.%m.%Y")

        # Steuern aus "Taxes paid / Tax Funds"-Bereich (Seite 2)
        section_match = re.search(
            r'Taxes paid / Tax Funds(.*?)Purchases taken into account',
            pdf_text,
            re.S
        )
        if section_match:
            section_text = section_match.group(1)
            minus_vals = re.findall(r'([\d\.,]+)\s*-\s*', section_text)
            if len(minus_vals) >= 3:
                # letzte 3 negativen Werte = Steuern
                cg = parse_number(minus_vals[-3])
                church = parse_number(minus_vals[-2])
                soli = parse_number(minus_vals[-1])
                trade_data['Capital Gains Tax'] = cg
                trade_data['Church Tax'] = church
                trade_data['Solidarity Surcharge'] = soli
                trade_data['Taxes Sum'] = cg + church + soli

        return trade_data

def find_next_empty_row_in_column(sheet, column_letter):
    """Find the next empty row in the specified column that has actual data."""
    for row in range(1, sheet.max_row + 1):
        if not sheet[f'{column_letter}{row}'].value:  # Check if the cell in column A is empty
            return row
    return sheet.max_row + 1  # If no empty row is found, return the next available row


def apply_date_format_to_column(sheet, column_letter):
    """Ensure the entire column has the same date format applied."""
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
    
    # If the style already exists, don't recreate it
    if "date_style" not in sheet.parent.named_styles:
        sheet.parent.add_named_style(date_style)

    for row in range(2, sheet.max_row + 1):  # Assuming row 1 is headers
        if isinstance(sheet[f'{column_letter}{row}'].value, datetime):
            sheet[f'{column_letter}{row}'].style = date_style

def choose_buy_row_interactively(sheet, source):
    """
    Zeigt offene bzw. relevante Buys an und lässt dich per a-z einen auswählen.
    source == 'cost_info': nur Zeilen ohne Sell-Datum (F leer)
    source == 'contract_note': alle Trades (auch schon geschlossene), damit du z.B. nur Steuern nachtragen kannst.
    """
    candidates = []
    for row in range(2, sheet.max_row + 1):
        name = sheet[f'A{row}'].value
        if not name:
            continue

        # Für Cost-Info nur offene Buys anzeigen
        if source == 'cost_info' and sheet[f'F{row}'].value:
            continue

        candidates.append(row)

    if not candidates:
        print("Keine passenden Buy-Kandidaten gefunden, es gibt keine offenen/geeigneten Buys.")
        return None

    letters = "abcdefghijklmnopqrstuvwxyz"
    max_entries = min(len(candidates), len(letters))

    print("\nKein automatisches Match gefunden. Wähle den passenden Buy-Eintrag:")
    for idx in range(max_entries):
        row = candidates[idx]
        letter = letters[idx]
        name = sheet[f'A{row}'].value
        buy_date = sheet[f'B{row}'].value
        qty = sheet[f'C{row}'].value
        print(f"  {letter}) Zeile {row}: {name} | Kaufdatum: {buy_date} | Menge: {qty}")

    if len(candidates) > max_entries:
        print(f"  ... ({len(candidates) - max_entries} weitere nicht angezeigt)")

    print("  (Enter leer lassen, um keinen Match auszuwählen)")

    while True:
        choice = input("Deine Auswahl (a-z oder leer für keinen Match): ").strip().lower()
        if choice == '':
            return None
        if len(choice) == 1 and choice in letters[:max_entries]:
            idx = letters.index(choice)
            return candidates[idx]
        print("Ungültige Eingabe, bitte a-z oder Enter.")


def update_excel(trade_data):
    # Read the existing Excel file or create a new DataFrame
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active
    else:
        print("Excel not found")
        return False  # ohne Excel kein Erfolg

    trade_type = trade_data['Trade Type']
    asset_name = trade_data['Asset Name']
    source = trade_data.get('Source', 'cost_info')  # 'cost_info' oder 'contract_note'

    # Ensure date format is applied correctly
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

    # Check if the date_style already exists to avoid the ValueError
    if "date_style" not in workbook.named_styles:
        workbook.add_named_style(date_style)

    if trade_type == 'Buy':
        next_row = find_next_empty_row_in_column(sheet, 'A')

        # Insert the new trade data into the predefined table
        sheet[f'A{next_row}'] = trade_data['Asset Name']
        buy_date = trade_data['Date']
        if isinstance(buy_date, str):
            buy_date = datetime.strptime(buy_date, "%d.%m.%Y").date()
        sheet[f'B{next_row}'] = buy_date
        sheet[f'B{next_row}'].style = "date_style"
        sheet[f'C{next_row}'] = trade_data['Quantity']
        sheet[f'D{next_row}'] = trade_data['Price per Unit']
        # sheet[f'K{next_row}'] = trade_data['Fees']

    elif trade_type == 'Sell':
        # Cost-Info-Sell: nur Zeilen ohne Sell-Datum
        # Contract-Note-Sell: auch Trades mit gesetztem Sell-Datum (nur Steuern nachtragen)
        require_empty_sell_date = (source == 'cost_info')

        match_row = find_best_matching_row(
            sheet,
            asset_name,
            require_empty_sell_date=require_empty_sell_date,
            min_score=0.5
        )

        if not match_row:
            print(f"No matching Buy entry found for asset {asset_name}.")
            # Interaktive Auswahl aller relevanten Buys
            manual_row = choose_buy_row_interactively(sheet, source)
            if manual_row is None:
                print("Kein manueller Match ausgewählt. Trade wird NICHT importiert.")
                workbook.save(EXCEL_FILE)
                return False
            match_row = manual_row

        # Sell-Daten eintragen, aber bei Contract-Notes nur überschreiben, wenn noch nichts da ist
        sell_date = trade_data.get('Date')
        if sell_date and (source == 'cost_info' or not sheet[f'F{match_row}'].value):
            if isinstance(sell_date, str):
                sell_date = datetime.strptime(sell_date, "%d.%m.%Y").date()
            sheet[f'F{match_row}'] = sell_date
            sheet[f'F{match_row}'].style = "date_style"

        if 'Quantity' in trade_data and (source == 'cost_info' or not sheet[f'G{match_row}'].value):
            sheet[f'G{match_row}'] = trade_data['Quantity']

        if 'Price per Unit' in trade_data and (source == 'cost_info' or not sheet[f'H{match_row}'].value):
            sheet[f'H{match_row}'] = trade_data['Price per Unit']

        # --- Steuern aus Contract Notes ---
        if 'Capital Gains Tax' in trade_data:
            sheet[f'Y{match_row}'] = trade_data['Capital Gains Tax']

        if 'Solidarity Surcharge' in trade_data:
            sheet[f'Z{match_row}'] = trade_data['Solidarity Surcharge']

        if 'Church Tax' in trade_data:
            sheet[f'AA{match_row}'] = trade_data['Church Tax']

        # Order Fees könntest du später hier einbauen

    # Gemeinsames Save & Erfolgsmeldung
    workbook.save(EXCEL_FILE)
    print(f"Excel file updated successfully with trade data for {trade_data['Asset Name']}.")
    return True

def process_pdfs():
    for filename in os.listdir(PDF_FOLDER):
        if not filename.lower().endswith('.pdf'):
            continue

        pdf_path = os.path.join(PDF_FOLDER, filename)
        print(f"Processing {filename}...")

        pdf_text = extract_text(pdf_path)

        trade_data = None

        # 1. Cost-Information?
        if 'Ex-Ante cost information' in pdf_text:
            trade_data = extract_cost_info_trade(pdf_text)
        else:
            # 2. Contract Note (Scalable oder Baader)?
            trade_data = extract_contract_note_trade(pdf_text)

        if trade_data:
            success = update_excel(trade_data)
            if success:
                os.remove(pdf_path)
                print(f"{filename} processed and deleted.")
            else:
                print(f"{filename} wurde NICHT gelöscht, weil der Trade nicht importiert wurde.")
        else:
            print(f"Failed to extract data from {filename}.")


if __name__ == "__main__":
    process_pdfs()
