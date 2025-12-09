import os
import re
import pandas as pd
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from datetime import datetime

# --- KONFIGURATION ---
# Pfad zum Ordner, in dem das Skript liegt
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_FOLDER = os.path.join(BASE_DIR, 'pdfs')
EXCEL_FILE = os.path.join(BASE_DIR, 'Trading.xlsx')

# ---------------------------------------------------------
# HILFSFUNKTIONEN
# ---------------------------------------------------------

def parse_german_float(text):
    """Wandelt deutsche (1.234,56) oder englische Zahlen in float um."""
    if not text:
        return 0.0
    # Bereinigen
    text = text.replace('EUR', '').replace('USD', '').replace('€', '').replace('$', '').strip()
    
    # Fall: 1.234,56 (Deutsch) -> Punkt entfernen, Komma zu Punkt
    if ',' in text and '.' in text:
        if text.find('.') < text.find(','):
            text = text.replace('.', '').replace(',', '.')
        else:
            text = text.replace(',', '') # Englisch (1,200.50)
    elif ',' in text:
        text = text.replace(',', '.')
    
    try:
        return float(text)
    except ValueError:
        return 0.0

def are_assets_similar(name1, name2):
    """Vergleicht zwei Asset-Namen tolerant (Intelligentes Matching)."""
    if not name1 or not name2:
        return False
    
    s1 = str(name1).lower()
    s2 = str(name2).lower()
    
    if s1 == s2: return True

    def clean_and_tokenize(text):
        text = re.sub(r'\d{2}\.\d{2}\.\d{2,4}', '', text) # Datum weg
        text = re.sub(r'[\$€]', '', text)
        text = text.replace('eur', '').replace('usd', '').replace('pc.', '').replace('stk.', '')
        text = text.replace(',', '.')
        tokens = text.split()
        cleaned = set()
        for t in tokens:
            try:
                # Zahlen normalisieren (200.00 -> 200)
                val = float(t)
                t = f"{val:.4f}".rstrip('0').rstrip('.')
            except ValueError:
                pass
            if len(t) > 1 or t.isdigit():
                cleaned.add(t)
        return cleaned

    tokens1 = clean_and_tokenize(s1)
    tokens2 = clean_and_tokenize(s2)
    
    common = tokens1.intersection(tokens2)
    min_len = min(len(tokens1), len(tokens2))
    if min_len == 0: return False
    
    # Erlaube 1 abweichendes Wort bei längeren Namen
    threshold = min_len if min_len < 3 else min_len - 1
    
    return len(common) >= threshold

# ---------------------------------------------------------
# PARSER FUNKTIONEN (PDFs)
# ---------------------------------------------------------

def extract_trade_info_transaction_statement(pdf_text: str):
    """
    Parser für ALTE Baader/Scalable 'Transaction Statement'.
    Hier stehen Steuern oft detailliert (Kapitalertragsteuer etc.).
    """
    data = {"Source": "transaction_statement", "Taxes": 0.0, "Fee": 0.0}
    
    # Typ (Sale/Purchase)
    if re.search(r'Transaction Statement:\s*Sale', pdf_text, re.IGNORECASE):
        data["Trade Type"] = "Sell"
    elif re.search(r'Transaction Statement:\s*(Purchase|Buy)', pdf_text, re.IGNORECASE):
        data["Trade Type"] = "Buy"
    else:
        return None

    # Datum
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', pdf_text)
    if date_match:
        y, m, d = date_match.group(1).split('-')
        data["Date"] = f"{d}.{m}.{y}"
    else:
        return None

    # Menge
    units_match = re.search(r'Units\s+([\d\.,]+)', pdf_text)
    if units_match:
        data["Quantity"] = parse_german_float(units_match.group(1))
    else:
        return None

    # Asset Name (Heuristiken wie im alten Code)
    lines = pdf_text.splitlines()
    asset_name = None
    start_idx = -1
    
    # Anker suchen (Price EUR oder ISIN)
    for i, line in enumerate(lines):
        if "Price" in line and "EUR" in line:
            start_idx = i
            break
    if start_idx == -1:
        for i, line in enumerate(lines):
            if "ISIN:" in line:
                start_idx = i
                break

    if start_idx != -1:
        for i in range(start_idx + 1, min(len(lines), start_idx + 10)):
            line = lines[i].strip()
            if not line: continue
            bad_words = ["Execution Venue", "Market Value", "Amount", "Order", "Account", "Baader", "Client", "Portfolio"]
            if any(bw.lower() in line.lower() for bw in bad_words): continue
            if re.match(r'^[\d\.,\s:-]+$', line): continue 
            if re.search(r'\d{4}-\d{2}-\d{2}', line): continue
            if "Price" in line and "EUR" in line: continue
            
            asset_name = line
            break
            
    if not asset_name:
        # Fallback: Suche im Quantity Block
        q_idx = pdf_text.find("Quantity")
        if q_idx != -1:
            snippet = pdf_text[q_idx:q_idx+400]
            for line in snippet.splitlines():
                if "Execution Venue" in line: continue
                if len(line) > 3 and not any(x in line for x in ["Quantity", "Units", "Price", "2025-", "Date"]):
                     asset_name = line.strip()
                     break

    data["Asset Name"] = asset_name if asset_name else "Unknown Asset"

    # Preis
    price_match = re.search(r'Price\s+EUR\s+([\d\.,]+)', pdf_text)
    if price_match:
        data["Price per Unit"] = parse_german_float(price_match.group(1))
    else:
        return None

    # Steuern (Summieren)
    taxes = 0.0
    tax_keywords = ["German flat rate tax", "Solidarity surcharge", "Church tax", "Kapitalertragsteuer", "Soli"]
    for line in lines:
        for key in tax_keywords:
            if key in line:
                # Suche Zahl am Ende der Zeile oder nach EUR
                m = re.search(r'([\d\.,]+)\s*-?$', line.strip())
                if not m: m = re.search(r'EUR\s+([\d\.,]+)', line)
                if m: taxes += parse_german_float(m.group(1))
    
    data["Taxes"] = taxes
    return data

def extract_trade_info_contract_note(pdf_text: str):
    """
    Parser für NEUE Scalable 'Contract Note'.
    Hier stehen Steuern oft als "Taxes" Block weiter unten.
    """
    data = {"Source": "contract_note", "Taxes": 0.0, "Fee": 0.0}

    # Typ & Asset (Erste Zeile: "Sell ... Asset Name")
    # Sucht nach Zeilenstart mit Buy/Sell gefolgt von Text
    head_match = re.search(r'^(Buy|Sell)\s+(?:[\d\.,]+(?:\s*pc\.|\s*Stk\.)?)?\s*(.+)$', pdf_text, re.MULTILINE)
    
    if head_match:
        data["Trade Type"] = head_match.group(1).capitalize()
        # Manchmal ist die Menge im Namen drin, manchmal davor. Hier vereinfacht:
        raw_name = head_match.group(2).strip()
        # Falls die Menge noch am Anfang des Namens klebt (z.B. "72.00 pc. Microsoft")
        name_cleanup = re.search(r'^[\d\.,]+\s*(?:pc\.|Stk\.)\s+(.+)', raw_name)
        if name_cleanup:
            data["Asset Name"] = name_cleanup.group(1).strip()
        else:
            data["Asset Name"] = raw_name
    else:
        return None

    # Datum
    date_match = re.search(r'(?:Execution|Date)\s+(\d{2}\.\d{2}\.\d{4})', pdf_text)
    if date_match: 
        data["Date"] = date_match.group(1)
    else: 
        return None

    # Menge & Preis
    # Pattern: "72.00 pc. 1.74 EUR"
    qp_match = re.search(r'([\d\.,]+)\s*(?:pc\.|Stk\.)\s+([\d\.,]+)\s*EUR', pdf_text)
    if qp_match:
        data["Quantity"] = parse_german_float(qp_match.group(1))
        data["Price per Unit"] = parse_german_float(qp_match.group(2))
    else: 
        return None

    # Gebühren (Order fees)
    # Sucht nach "Order fees" und dann einer Zahl, auch über Zeilenumbrüche hinweg
    fee_match = re.search(r'Order fees\s*[\r\n]*\s*([-\d\.,]+)\s*EUR', pdf_text, re.MULTILINE)
    if fee_match: 
        data["Fee"] = abs(parse_german_float(fee_match.group(1)))

    # Steuern (Taxes)
    # Sucht nach "Taxes" gefolgt von Zahl, erlaubt Newlines/Whitespace
    # Wichtig: Scalable schreibt oft negativ (-1.26 EUR), wir wollen positiv für Excel
    tax_match = re.search(r'Taxes\s*[\r\n]*\s*([-\d\.,]+)\s*EUR', pdf_text, re.MULTILINE)
    if tax_match: 
        data["Taxes"] = abs(parse_german_float(tax_match.group(1)))

    return data

def extract_trade_info_cost_information(pdf_text: str):
    """Parser für 'Ex-Ante cost information' (meist keine echten Steuern, nur Gebührenschätzung)."""
    data = {"Source": "cost_information", "Taxes": 0.0, "Fee": 0.0}
    
    if "Order Buy" in pdf_text or "Order\nBuy" in pdf_text: data["Trade Type"] = "Buy"
    elif "Order Sell" in pdf_text or "Order\nSell" in pdf_text: data["Trade Type"] = "Sell"
    else: return None

    asset_match = re.search(r'Ex-Ante cost information\s*\n\s*(.+)', pdf_text)
    data["Asset Name"] = asset_match.group(1).strip() if asset_match else "Unknown"

    date_match = re.search(r'Date\s+(\d{2}\.\d{2}\.\d{4})', pdf_text)
    if date_match: data["Date"] = date_match.group(1)
    else: return None

    qty_match = re.search(r'Quantity\s+([\d\.,]+)', pdf_text)
    if qty_match: data["Quantity"] = parse_german_float(qty_match.group(1))
    else: return None

    # Preis schätzen aus "Est. order amount"
    amt_match = re.search(r'Est\. order amount\s+([\d\.,]+)\s*€', pdf_text)
    if amt_match and data["Quantity"] > 0:
        total = parse_german_float(amt_match.group(1))
        data["Price per Unit"] = total / data["Quantity"]
    else:
        data["Price per Unit"] = 0.0
        
    # Service Charges als Fee annehmen
    fee_match = re.search(r'Service charges\s+([\d\.,]+)\s*EUR', pdf_text)
    if fee_match:
        data["Fee"] = parse_german_float(fee_match.group(1))

    return data

def extract_trade_info(pdf_path):
    try:
        text = extract_text(pdf_path)
    except Exception as e:
        print(f"Fehler beim Lesen von {pdf_path}: {e}")
        return None

    if "Transaction Statement" in text:
        return extract_trade_info_transaction_statement(text)
    elif "Contract note" in text or "Abrechnung" in text:
        return extract_trade_info_contract_note(text)
    elif "Ex-Ante cost information" in text or "Kostentransparenz" in text:
        return extract_trade_info_cost_information(text)
    else:
        print(f"Unbekanntes PDF Format: {os.path.basename(pdf_path)}")
        return None

# ---------------------------------------------------------
# EXCEL HANDLING (NUR ROHDATEN SCHREIBEN)
# ---------------------------------------------------------

def ensure_columns_exist(sheet):
    """
    Liest die Header aus Zeile 1 und gibt ein Mapping {Name: Buchstabe} zurück.
    Erstellt KEINE neuen Spalten mehr, um Layout nicht zu zerschießen.
    """
    headers = {}
    for cell in sheet[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column_letter
    return headers

def find_buy_row(sheet, asset_name, quantity):
    """Sucht die Kauf-Zeile."""
    for row in range(2, sheet.max_row + 1):
        name_cell = sheet[f"A{row}"]
        sell_date_cell = sheet[f"F{row}"]
        
        # Nur Zeilen ohne Verkaufsdatum betrachten
        if sell_date_cell.value:
            continue
            
        excel_name = str(name_cell.value).strip() if name_cell.value else ""
        
        if are_assets_similar(excel_name, asset_name):
            # Optional: Check Quantity match für mehr Sicherheit
            # qty_cell = sheet[f"C{row}"]
            # if qty_cell.value and abs(float(qty_cell.value) - quantity) < 0.001:
            return row
            
    return None

def write_excel(all_trades):
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel-Datei nicht gefunden unter: {EXCEL_FILE}")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    header_map = ensure_columns_exist(ws)
    
    # Wichtige Spalten identifizieren
    col_taxes = header_map.get("Taxes")
    col_fee = header_map.get("Trading Fee")
    
    # Debug Info
    if not col_taxes: print("ACHTUNG: Spalte 'Taxes' nicht gefunden. Steuern werden nicht eingetragen.")
    if not col_fee: print("ACHTUNG: Spalte 'Trading Fee' nicht gefunden. Gebühren werden nicht eingetragen.")

    date_style = NamedStyle(name="date_style_final_v3", number_format="DD.MM.YYYY")
    if "date_style_final_v3" not in wb.named_styles:
        wb.add_named_style(date_style)

    for trade in all_trades:
        asset = trade["Asset Name"]
        date_obj = datetime.strptime(trade["Date"], "%d.%m.%Y")
        qty = trade["Quantity"]
        price = trade["Price per Unit"]
        taxes = trade.get("Taxes", 0.0)
        fee = trade.get("Fee", 0.0)

        if trade["Trade Type"] == "Buy":
            # Neue Zeile anhängen
            row = ws.max_row + 1
            if not ws[f"A{ws.max_row}"].value and ws.max_row > 1:
                row = ws.max_row
            
            # NUR INPUTS SCHREIBEN (keine Formeln überschreiben in E, etc.)
            ws[f"A{row}"] = asset
            ws[f"B{row}"] = date_obj
            ws[f"B{row}"].style = "date_style_final_v3"
            ws[f"C{row}"] = qty
            ws[f"D{row}"] = price
            
            # Wenn du Formeln hast (Spalte E), werden diese NICHT angefasst.
            # Falls du eine "intelligente Tabelle" (ListObject) nutzt, füllt Excel E automatisch.
            # Falls nicht, musst du die Formel manuell runterziehen.
            print(f"BUY eingetragen: {asset} (Zeile {row})")

        elif trade["Trade Type"] == "Sell":
            row = find_buy_row(ws, asset, qty)
            if row:
                # NUR INPUTS SCHREIBEN
                ws[f"F{row}"] = date_obj
                ws[f"F{row}"].style = "date_style_final_v3"
                ws[f"G{row}"] = qty
                ws[f"H{row}"] = price
                
                # Gebühren und Steuern eintragen, wo sie hingehören
                if col_fee:
                    ws[f"{col_fee}{row}"] = fee
                
                if col_taxes:
                    ws[f"{col_taxes}{row}"] = taxes

                # KEINE Formeln schreiben (K, M, N werden ignoriert)
                print(f"SELL gematched: '{asset}' in Zeile {row}. Tax: {taxes}, Fee: {fee}")
            else:
                print(f"WARNUNG: Kein Match für Sell '{asset}' gefunden.")

    try:
        wb.save(EXCEL_FILE)
        print("Excel Update abgeschlossen.")
    except PermissionError:
        print("FEHLER: Bitte Excel-Datei schließen!")

def process_pdfs():
    if not os.path.exists(PDF_FOLDER):
        os.makedirs(PDF_FOLDER)
        return

    files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    if not files:
        print("Keine PDFs gefunden.")
        return

    merged_trades = {}

    for filename in files:
        path = os.path.join(PDF_FOLDER, filename)
        # print(f"Lese {filename}...")
        trade_data = extract_trade_info(path)
        
        if trade_data:
            trade_data['_filename'] = filename
            # Merging Logik (z.B. Cost Info vs Contract Note)
            key = (trade_data['Date'], trade_data['Trade Type'], trade_data['Quantity'])
            
            if key in merged_trades:
                existing = merged_trades[key]
                # Contract Note gewinnt immer (weil genauer)
                if trade_data['Source'] == 'contract_note':
                    merged_trades[key] = trade_data
                elif existing['Source'] == 'contract_note':
                    pass
                # Transaction Statement ist auch besser als Cost Info
                elif trade_data['Source'] == 'transaction_statement' and existing['Source'] == 'cost_information':
                    merged_trades[key] = trade_data
            else:
                merged_trades[key] = trade_data
            
            try: os.remove(path) 
            except: pass

    final_list = list(merged_trades.values())
    
    # Sortierung: Erst Käufe, dann Verkäufe
    def sort_algo(t):
        try:
            d, m, y = t['Date'].split('.')
            type_prio = 0 if t['Trade Type'] == 'Buy' else 1
            return (int(y)*10000 + int(m)*100 + int(d), type_prio)
        except:
            return (0, 0)

    final_list.sort(key=sort_algo)
    
    if final_list:
        write_excel(final_list)
    else:
        print("Keine verarbeitbaren Trades gefunden.")

if __name__ == "__main__":
    process_pdfs()