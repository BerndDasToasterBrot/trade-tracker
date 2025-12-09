import os
import re
import pandas as pd
from pdfminer.high_level import extract_text
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from datetime import datetime

# --- KONFIGURATION ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PDF_FOLDER = os.path.join(BASE_DIR, 'pdfs')
EXCEL_FILE = os.path.join(BASE_DIR, 'Trading.xlsx')

# ---------------------------------------------------------
# HILFSFUNKTIONEN
# ---------------------------------------------------------

def debug_print(msg):
    """Zentrale Debug-Funktion."""
    print(f"[DEBUG] {msg}")

def parse_german_float(text):
    if not text: return 0.0
    original = text
    # Entferne Währungen und Anführungszeichen/Kommas aus CSV-Artefakten
    text = text.replace('EUR', '').replace('USD', '').replace('€', '').replace('$', '').replace('"', '').strip()
    
    if ',' in text and '.' in text:
        if text.find('.') < text.find(','):
            text = text.replace('.', '').replace(',', '.')
        else:
            text = text.replace(',', '')
    elif ',' in text:
        text = text.replace(',', '.')
    
    try:
        val = float(text)
        # debug_print(f"Parsed Float: '{original}' -> {val}")
        return val
    except ValueError:
        debug_print(f"ERROR: Konnte '{original}' nicht als Zahl parsen.")
        return 0.0

def are_assets_similar(name1, name2):
    if not name1 or not name2: return False
    s1 = str(name1).lower()
    s2 = str(name2).lower()
    if s1 == s2: return True

    def clean_and_tokenize(text):
        text = re.sub(r'\d{2}\.\d{2}\.\d{2,4}', '', text)
        text = re.sub(r'[\$€]', '', text)
        text = text.replace('eur', '').replace('usd', '').replace('pc.', '').replace('stk.', '')
        text = text.replace(',', '.')
        text = text.replace('"', '') 
        tokens = text.split()
        cleaned = set()
        for t in tokens:
            try:
                val = float(t)
                t = f"{val:.4f}".rstrip('0').rstrip('.')
            except ValueError: pass
            if len(t) > 1 or t.isdigit():
                cleaned.add(t)
        return cleaned

    tokens1 = clean_and_tokenize(s1)
    tokens2 = clean_and_tokenize(s2)
    common = tokens1.intersection(tokens2)
    min_len = min(len(tokens1), len(tokens2))
    if min_len == 0: return False
    threshold = min_len if min_len < 3 else min_len - 1
    
    is_match = len(common) >= threshold
    if is_match:
        debug_print(f"Matching erfolgreich: '{name1}' == '{name2}'")
    # else:
    #    debug_print(f"Kein Match: '{name1}' vs '{name2}'")
    return is_match

# ---------------------------------------------------------
# PARSER FUNKTIONEN (PDFs)
# ---------------------------------------------------------

def extract_trade_info_transaction_statement(pdf_text: str):
    debug_print("Start Parser: Transaction Statement")
    data = {"Source": "transaction_statement", "Taxes": 0.0, "Fee": 0.0}
    
    if re.search(r'Transaction Statement:\s*Sale', pdf_text, re.IGNORECASE):
        data["Trade Type"] = "Sell"
    elif re.search(r'Transaction Statement:\s*(Purchase|Buy)', pdf_text, re.IGNORECASE):
        data["Trade Type"] = "Buy"
    else: 
        debug_print("Trade Type nicht gefunden (Transaction Statement).")
        return None

    date_match = re.search(r'(\d{4}-\d{2}-\d{2})', pdf_text)
    if date_match:
        y, m, d = date_match.group(1).split('-')
        data["Date"] = f"{d}.{m}.{y}"
    else: 
        debug_print("Datum nicht gefunden.")
        return None

    units_match = re.search(r'Units\s+([\d\.,]+)', pdf_text)
    if units_match:
        data["Quantity"] = parse_german_float(units_match.group(1))
    else: 
        debug_print("Units nicht gefunden.")
        return None

    # Name Extraction 
    lines = pdf_text.splitlines()
    asset_name = None
    start_idx = -1
    for i, line in enumerate(lines):
        if "Price" in line and "EUR" in line:
            start_idx = i
            break
    if start_idx == -1:
        for i, line in enumerate(lines):
            if "ISIN:" in line:
                start_idx = i
                break

    if start_idx != -1:
        for i in range(start_idx + 1, min(len(lines), start_idx + 15)):
            line = lines[i].strip()
            if not line: continue
            bad_words = ["Execution Venue", "Market Value", "Amount", "Order", "Account", 
                         "Baader", "Client", "Portfolio", "WKN", "ISIN", "UniCredit", 
                         "Bank", "Sitz", "Munich", "München", "Tax", "Reference"]
            if any(bw.lower() in line.lower() for bw in bad_words): continue
            if re.match(r'^[\d\.,\s:-]+$', line): continue 
            if re.search(r'\d{4}-\d{2}-\d{2}', line): continue
            if "Price" in line and "EUR" in line: continue
            asset_name = line
            break
            
    if not asset_name:
        q_idx = pdf_text.find("Quantity")
        if q_idx != -1:
            snippet = pdf_text[q_idx:q_idx+400]
            for line in snippet.splitlines():
                if "Execution Venue" in line: continue
                if len(line) > 3 and not any(x in line for x in ["Quantity", "Units", "Price", "2025-", "Date"]):
                     asset_name = line.strip()
                     break

    data["Asset Name"] = asset_name if asset_name else "Unknown Asset"
    debug_print(f"Gefundener Name: {data['Asset Name']}")

    price_match = re.search(r'Price\s+EUR\s+([\d\.,]+)', pdf_text)
    if price_match: data["Price per Unit"] = parse_german_float(price_match.group(1))
    else: 
        debug_print("Price nicht gefunden.")
        return None

    taxes = 0.0
    tax_keywords = ["German flat rate tax", "Solidarity surcharge", "Church tax", "Kapitalertragsteuer", "Soli"]
    for line in lines:
        for key in tax_keywords:
            if key in line:
                m = re.search(r'([\d\.,]+)\s*-?$', line.strip())
                if not m: m = re.search(r'EUR\s+([\d\.,]+)', line)
                if m: taxes += parse_german_float(m.group(1))
    data["Taxes"] = taxes
    return data

def extract_trade_info_contract_note(pdf_text: str):
    debug_print("Start Parser: Contract Note")
    data = {"Source": "contract_note", "Taxes": 0.0, "Fee": 0.0}

    head_match = re.search(r'^(Buy|Sell)\s+(?:[\d\.,]+(?:\s*pc\.|\s*Stk\.)?)?\s*(.+)$', pdf_text, re.MULTILINE)
    if head_match:
        data["Trade Type"] = head_match.group(1).capitalize()
        raw_name = head_match.group(2).strip()
        name_cleanup = re.search(r'^[\d\.,]+\s*(?:pc\.|Stk\.)\s+(.+)', raw_name)
        if name_cleanup: data["Asset Name"] = name_cleanup.group(1).strip()
        else: data["Asset Name"] = raw_name
        debug_print(f"Gefundener Name: {data['Asset Name']}")
    else: 
        debug_print("Header (Buy/Sell + Name) nicht gefunden.")
        return None

    date_match = re.search(r'(?:Execution|Date)\s+(\d{2}\.\d{2}\.\d{4})', pdf_text)
    if date_match: data["Date"] = date_match.group(1)
    else: 
        debug_print("Datum nicht gefunden.")
        return None

    qp_match = re.search(r'([\d\.,]+)\s*(?:pc\.|Stk\.)\s+([\d\.,]+)\s*EUR', pdf_text)
    if qp_match:
        data["Quantity"] = parse_german_float(qp_match.group(1))
        data["Price per Unit"] = parse_german_float(qp_match.group(2))
    else: 
        debug_print("Quantity/Price Block nicht gefunden.")
        return None

    fee_match = re.search(r'Order fees\s*[\r\n]*\s*([-\d\.,]+)\s*EUR', pdf_text, re.MULTILINE)
    if fee_match: data["Fee"] = abs(parse_german_float(fee_match.group(1)))

    tax_match = re.search(r'Taxes\s*[\r\n]*\s*([-\d\.,]+)\s*EUR', pdf_text, re.MULTILINE)
    if tax_match: data["Taxes"] = abs(parse_german_float(tax_match.group(1)))

    return data

def extract_trade_info_cost_information(pdf_text: str):
    debug_print("Start Parser: Ex-Ante Cost Information")
    data = {"Source": "cost_information", "Taxes": 0.0, "Fee": 0.0}
    
    # 1. Trade Type
    # Regex: Order["\s,\n]*(Buy|Sell) erlaubt Anführungszeichen, Newlines, Kommas dazwischen
    type_match = re.search(r'Order["\s,\n]*(Buy|Sell)', pdf_text, re.IGNORECASE)
    if type_match:
        data["Trade Type"] = type_match.group(1).capitalize()
        debug_print(f"Trade Type gefunden: {data['Trade Type']}")
    else: 
        debug_print("Trade Type (Order Buy/Sell) NICHT gefunden. Checke PDF-Inhalt:")
        debug_print(pdf_text[:200].replace('\n', '\\n'))
        return None

    # 2. Asset Name
    # Sucht nach dem Header "Ex-Ante cost information" und nimmt die nächste Zeile (oder Text danach)
    asset_match = re.search(r'Ex-Ante cost information\s+([^\r\n]+)', pdf_text)
    if not asset_match:
         asset_match = re.search(r'Ex-Ante cost information\s*\n\s*([^\r\n]+)', pdf_text)
    
    if asset_match:
        data["Asset Name"] = asset_match.group(1).strip()
        debug_print(f"Name gefunden: {data['Asset Name']}")
    else:
        data["Asset Name"] = "Unknown"
        debug_print("WARNUNG: Name nicht gefunden.")

    # 3. Datum
    date_match = re.search(r'Date["\s,\n]*(\d{2}\.\d{2}\.\d{4})', pdf_text)
    if date_match: 
        data["Date"] = date_match.group(1)
        debug_print(f"Date gefunden: {data['Date']}")
    else: 
        debug_print("Datum NICHT gefunden.")
        return None

    # 4. Menge
    qty_match = re.search(r'Quantity["\s,\n]*([\d\.,]+)', pdf_text)
    if qty_match: 
        data["Quantity"] = parse_german_float(qty_match.group(1))
    else: 
        debug_print("Quantity NICHT gefunden.")
        return None

    # 5. Preis (berechnet aus Est. Amount / Quantity)
    amt_match = re.search(r'Est\. order amount["\s,\n]*([\d\.,]+)', pdf_text)
    if amt_match and data["Quantity"] > 0:
        total = parse_german_float(amt_match.group(1))
        data["Price per Unit"] = total / data["Quantity"]
    else:
        debug_print("Est. order amount nicht gefunden oder Quantity 0 -> Preis = 0")
        data["Price per Unit"] = 0.0
        
    fee_match = re.search(r'Service charges["\s,\n]*([\d\.,]+)', pdf_text)
    if fee_match: data["Fee"] = parse_german_float(fee_match.group(1))

    return data

def extract_trade_info(pdf_path):
    print(f"\nVerarbeite Datei: {os.path.basename(pdf_path)}")
    try: text = extract_text(pdf_path)
    except Exception as e:
        debug_print(f"ERROR: Fehler beim Lesen: {e}")
        return None

    if "Transaction Statement" in text:
        return extract_trade_info_transaction_statement(text)
    elif "Contract note" in text or "Abrechnung" in text:
        return extract_trade_info_contract_note(text)
    elif "Ex-Ante cost information" in text or "Kostentransparenz" in text:
        return extract_trade_info_cost_information(text)
    else:
        debug_print(f"WARNUNG: Unbekanntes PDF Format.")
        debug_print("Erste 100 Zeichen des Textes:")
        debug_print(text[:100].replace('\n', '\\n'))
        return None

# ---------------------------------------------------------
# EXCEL HANDLING
# ---------------------------------------------------------

def ensure_columns_exist(sheet):
    headers = {}
    for cell in sheet[1]:
        if cell.value: headers[str(cell.value).strip()] = cell.column_letter
    return headers

def find_buy_row(sheet, asset_name, quantity):
    debug_print(f"Suche BUY-Zeile für '{asset_name}' (Qty: {quantity})")
    for row in range(2, sheet.max_row + 1):
        name_cell = sheet[f"A{row}"]
        sell_date_cell = sheet[f"F{row}"]
        
        # Nur Zeilen ohne Verkaufsdatum
        if sell_date_cell.value: continue 
            
        excel_name = str(name_cell.value).strip() if name_cell.value else ""
        if are_assets_similar(excel_name, asset_name):
            debug_print(f" -> TREFFER in Zeile {row}")
            return row
    debug_print(" -> KEIN TREFFER gefunden.")
    return None

def write_excel(all_trades):
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Excel-Datei nicht gefunden unter: {EXCEL_FILE}")
        return

    print(f"Öffne Excel: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    date_style = NamedStyle(name="date_style_final_v3", number_format="DD.MM.YYYY")
    if "date_style_final_v3" not in wb.named_styles: wb.add_named_style(date_style)

    header_map = ensure_columns_exist(ws)
    col_taxes = header_map.get("Taxes")
    col_fee = header_map.get("Trading Fee")

    for trade in all_trades:
        asset = trade["Asset Name"]
        date_obj = datetime.strptime(trade["Date"], "%d.%m.%Y")
        qty = trade["Quantity"]
        price = trade["Price per Unit"]
        taxes = trade.get("Taxes", 0.0)
        fee = trade.get("Fee", 0.0)

        if trade["Trade Type"] == "Buy":
            row = ws.max_row + 1
            if ws.max_row > 1 and not ws[f"A{ws.max_row}"].value:
                row = ws.max_row
            
            print(f"Eintrag BUY: {asset} (Zeile {row})")
            ws[f"A{row}"] = asset
            ws[f"B{row}"] = date_obj
            ws[f"B{row}"].style = "date_style_final_v3"
            ws[f"C{row}"] = qty
            ws[f"D{row}"] = price
            
            try: wb.save(EXCEL_FILE)
            except PermissionError: 
                print("CRITICAL ERROR: Excel ist offen! Bitte schließen.")
                return

        elif trade["Trade Type"] == "Sell":
            row = find_buy_row(ws, asset, qty)
            if row:
                print(f"Match SELL: '{asset}' in Zeile {row}")
                ws[f"F{row}"] = date_obj
                ws[f"F{row}"].style = "date_style_final_v3"
                ws[f"G{row}"] = qty
                ws[f"H{row}"] = price
                if col_fee: ws[f"{col_fee}{row}"] = fee
                if col_taxes: ws[f"{col_taxes}{row}"] = taxes
                
                try: wb.save(EXCEL_FILE)
                except PermissionError: 
                    print("CRITICAL ERROR: Excel ist offen! Bitte schließen.")
                    return
            else:
                print(f"WARNUNG: Kein Match für Sell '{asset}' gefunden.")

    print("\n--- FERTIG ---")

def process_pdfs():
    if not os.path.exists(PDF_FOLDER):
        os.makedirs(PDF_FOLDER)
        return

    files = [f for f in os.listdir(PDF_FOLDER) if f.lower().endswith('.pdf')]
    if not files:
        print("Keine PDFs gefunden.")
        return

    merged_trades = {}

    for filename in files:
        path = os.path.join(PDF_FOLDER, filename)
        trade_data = extract_trade_info(path)
        
        if trade_data:
            trade_data['_filename'] = filename
            key = (trade_data['Date'], trade_data['Trade Type'], trade_data['Quantity'])
            
            if key in merged_trades:
                existing = merged_trades[key]
                debug_print(f"Duplikat gefunden für {key}")
                debug_print(f"  Alt: {existing['Source']} ({existing['Asset Name']})")
                debug_print(f"  Neu: {trade_data['Source']} ({trade_data['Asset Name']})")
                
                # --- MERGE LOGIK ---
                # 1. Neues ist Contract Note (beste Daten)
                if trade_data['Source'] == 'contract_note':
                    if existing['Source'] == 'cost_information':
                        debug_print("  -> Nehme Contract Note Daten, behalte Cost Info Namen")
                        trade_data['Asset Name'] = existing['Asset Name']
                    merged_trades[key] = trade_data
                
                # 2. Altes ist Contract Note
                elif existing['Source'] == 'contract_note':
                    if trade_data['Source'] == 'cost_information':
                         debug_print("  -> Behalte Contract Note Daten, update Name von Cost Info")
                         existing['Asset Name'] = trade_data['Asset Name']
                         merged_trades[key] = existing
                
                # 3. Transaction vs Cost Info
                elif trade_data['Source'] == 'transaction_statement' and existing['Source'] == 'cost_information':
                     debug_print("  -> Nehme Transaction St. Daten, behalte Cost Info Namen")
                     trade_data['Asset Name'] = existing['Asset Name']
                     merged_trades[key] = trade_data
                
                # 4. Cost Info (neu) vs Transaction (alt)
                elif trade_data['Source'] == 'cost_information' and existing['Source'] == 'transaction_statement':
                     debug_print("  -> Behalte Transaction St. Daten, update Name von Cost Info")
                     existing['Asset Name'] = trade_data['Asset Name']
                     merged_trades[key] = existing

                else:
                    merged_trades[key] = trade_data
            else:
                merged_trades[key] = trade_data
            
            try: os.remove(path) 
            except: pass
        else:
            print(f"WARNUNG: Konnte {filename} nicht lesen.")

    final_list = list(merged_trades.values())
    
    # Sortierung: Datum -> Buy vor Sell
    def sort_algo(t):
        try:
            d, m, y = t['Date'].split('.')
            type_prio = 0 if t['Trade Type'] == 'Buy' else 1
            return (int(y)*10000 + int(m)*100 + int(d), type_prio)
        except: return (0, 0)

    final_list.sort(key=sort_algo)
    
    if final_list:
        write_excel(final_list)
    else:
        print("Keine verarbeitbaren Trades gefunden.")

if __name__ == "__main__":
    process_pdfs()